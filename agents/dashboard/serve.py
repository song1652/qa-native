"""
QA Agent Dashboard 서버
프로젝트 루트 또는 어디서든 실행 가능:
  python agents/dashboard/serve.py
"""
import json
import queue
import re
import sys
import threading
import time
import webbrowser
from datetime import datetime
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path

PORT = 8766
HERE = Path(__file__).parent                    # agents/dashboard/
PROJECT_ROOT = HERE.parent.parent               # qa-native/

sys.path.insert(0, str(PROJECT_ROOT / "scripts"))
# .venv site-packages를 sys.path에 추가 (시스템 Python으로 실행해도 패키지 사용 가능)
_venv_sp = PROJECT_ROOT / ".venv" / "lib"
if _venv_sp.exists():
    for _sp in _venv_sp.glob("python*/site-packages"):
        if str(_sp) not in sys.path:
            sys.path.insert(0, str(_sp))
from _python import PYTHON_EXE
DIALOG_PATH = PROJECT_ROOT / "agents" / "dialog.json"
STATE_PATH = PROJECT_ROOT / "state" / "pipeline.json"
TEAM_NOTES_PATH = PROJECT_ROOT / "agents" / "team_notes.md"
DISCUSS_PATH = PROJECT_ROOT / "state" / "discuss.json"
PENDING_IMPL_PATH = PROJECT_ROOT / "pending_impl.json"
PARALLEL_STATE_PATH = PROJECT_ROOT / "state" / "parallel.json"
GENERATED_DIR = PROJECT_ROOT / "tests" / "generated"
REPORTS_DIR = PROJECT_ROOT / "tests" / "reports"
QUICK_STATE_PATH = PROJECT_ROOT / "state" / "quick.json"
LOGS_DIR = PROJECT_ROOT / "logs"
LOGS_DIR.mkdir(exist_ok=True)
IMPORT_DIR = PROJECT_ROOT / "import"

# ── SSE 클라이언트 관리 ────────────────────────────────────────
_sse_clients: list[queue.Queue] = []
_sse_lock = threading.Lock()


def _sse_notify():
    """dialog.json 변경 시 모든 SSE 클라이언트에 알림."""
    with _sse_lock:
        dead = []
        for q in _sse_clients:
            try:
                q.put_nowait("update")
            except queue.Full:
                dead.append(q)
        for q in dead:
            _sse_clients.remove(q)


def _watch_files():
    """dialog.json / state/discuss.json mtime을 0.3초마다 감시."""
    watched = [DIALOG_PATH, DISCUSS_PATH, STATE_PATH, PARALLEL_STATE_PATH,
               QUICK_STATE_PATH]
    last_mtimes = {p: 0.0 for p in watched}
    while True:
        for p in watched:
            try:
                mtime = p.stat().st_mtime if p.exists() else 0.0
                if mtime != last_mtimes[p]:
                    last_mtimes[p] = mtime
                    _sse_notify()
            except Exception:
                pass
        time.sleep(0.3)


# 파일 감시 스레드 시작
threading.Thread(target=_watch_files, daemon=True).start()

TEAM_NOTES_HEADER = (
    "# 팀 결정 사항\n\n"
    "> **독자**: 심의 Agent — 팀 토론 결론 누적. 토론 시 중복 결론 방지 목적으로 참조.\n\n"
    "---\n"
)


def load_json(path: Path):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return None
    return None


def parse_conclusion_items(conclusion: str) -> list:
    """결론 마크다운을 투표 가능한 개별 항목으로 파싱."""
    items = []

    # 1) ### 소제목 파싱
    for m in re.finditer(r'^###\s+(.+?)\n([\s\S]*?)(?=^###\s|\Z)', conclusion, re.MULTILINE):
        title = m.group(1).strip()
        body  = re.sub(r'\n?---+\s*$', '', m.group(2)).strip()
        items.append({"id": len(items), "title": title, "text": body, "status": "pending"})

    # 2) 번호 목록 파싱 (1. **title**: body)
    for m in re.finditer(r'^\d+\.\s+(.+)$', conclusion, re.MULTILINE):
        text = m.group(1).strip()
        bold = re.match(r'\*\*(.+?)\*\*[:\s]*(.*)', text)
        title = bold.group(1).strip() if bold else text[:70]
        items.append({"id": len(items), "title": title, "text": text, "status": "pending"})

    # 3) fallback
    if not items:
        items.append({"id": 0, "title": "전체 결론", "text": conclusion, "status": "pending"})

    return items


def finalize_team_notes(discuss: dict):
    """승인된 항목만 team_notes.md에 덮어쓰기 + pending_impl.json 생성."""
    import datetime
    items = discuss.get("conclusion_items", [])
    approved = [i for i in items if i["status"] == "approved"]
    topic = discuss.get("topic", "")
    today = datetime.datetime.now().strftime("%Y-%m-%d")

    content = TEAM_NOTES_HEADER
    if approved:
        content += f"\n## {topic}\n> 결정일: {today}\n\n"
        for item in approved:
            content += f"### {item['title']}\n{item['text']}\n\n"
        content += "---\n"

    TEAM_NOTES_PATH.write_text(content, encoding="utf-8")

    # 구현 대기 파일 생성 → UserPromptSubmit 훅이 감지해 Claude에 주입
    if approved:
        pending = {
            "status": "pending",
            "topic": topic,
            "approved_at": datetime.datetime.now().isoformat(),
            "items": approved,
        }
        PENDING_IMPL_PATH.write_text(
            json.dumps(pending, ensure_ascii=False, indent=2), encoding="utf-8"
        )


def build_pipeline_state() -> dict:
    """단일 파이프라인 state/pipeline.json 반환."""
    return load_json(STATE_PATH) or {}


def build_batch_state() -> dict:
    """병렬 파이프라인 상태 + tests/generated/ 파일 목록 반환."""
    parallel = load_json(PARALLEL_STATE_PATH) or {}
    generated_files = []
    if GENERATED_DIR.exists():
        for group_dir in sorted(GENERATED_DIR.iterdir()):
            if group_dir.is_dir() and not group_dir.name.startswith("."):
                for f in sorted(group_dir.glob("*.py")):
                    if f.name not in ("conftest.py", "__init__.py"):
                        generated_files.append({
                            "group": group_dir.name,
                            "file": f.name,
                            "path": str(f.relative_to(PROJECT_ROOT)),
                            "size": f.stat().st_size,
                        })
    return {"parallel_state": parallel, "generated_files": generated_files}


PAGES_JSON = PROJECT_ROOT / "config" / "pages.json"
TESTCASES_DIR = PROJECT_ROOT / "testcases"


def list_pages() -> dict:
    """config/pages.json 반환."""
    return load_json(PAGES_JSON) or {}


def list_testcase_groups() -> list:
    """testcases/ 하위 폴더별 케이스 파일 목록."""
    if not TESTCASES_DIR.exists():
        return []
    groups = []
    for d in sorted(TESTCASES_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith("."):
            continue
        cases = sorted([f.name for f in d.glob("tc_*.md")])
        groups.append({"name": d.name, "cases": cases, "count": len(cases)})
    return groups


def list_generated_groups() -> list:
    """tests/generated/ 하위 그룹별 테스트 파일 목록 반환."""
    if not GENERATED_DIR.exists():
        return []
    groups = []
    for d in sorted(GENERATED_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith((".", "_")):
            continue
        files = sorted([
            f.name for f in d.glob("*.py")
            if f.name not in ("conftest.py", "__init__.py")
        ])
        if files:
            groups.append({
                "name": d.name,
                "file_count": len(files),
                "files": files,
            })
    return groups


def list_reports() -> list:
    """tests/reports/ 의 HTML 파일 목록 (최신순)."""
    if not REPORTS_DIR.exists():
        return []
    reports = []
    for f in sorted(REPORTS_DIR.glob("*.html"), key=lambda p: p.stat().st_mtime, reverse=True):
        reports.append({
            "name": f.name,
            "modified_at": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "size": f.stat().st_size,
        })
    return reports


def build_dialogs() -> dict:
    """팀 토론 대화 payload 반환 (dialog.json은 팀 토론 전용)."""
    full_dialog = load_json(DIALOG_PATH) or {"sessions": []}
    discuss_state = load_json(DISCUSS_PATH) or {}

    # step=discussed 이고 conclusion_items 없으면 자동 파싱
    if (discuss_state.get("step") == "discussed"
            and discuss_state.get("conclusion")
            and not discuss_state.get("conclusion_items")):
        discuss_state["conclusion_items"] = parse_conclusion_items(discuss_state["conclusion"])
        DISCUSS_PATH.write_text(
            json.dumps(discuss_state, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    all_sessions = full_dialog.get("sessions", [])
    team_sessions = [s for s in all_sessions if s.get("stage") == "team_discussion"]

    return {
        "team_sessions": team_sessions,
        "discuss_state": discuss_state,
    }


# ── Excel Import 유틸 ─────────────────────────────────────────
def _list_import_files() -> list:
    """import/ 폴더의 .xlsx 파일 목록."""
    if not IMPORT_DIR.exists():
        return []
    return sorted([f.name for f in IMPORT_DIR.glob("*.xlsx")])


def _detect_header_row(ws) -> int | None:
    """'Test\\nScenario ID' 패턴이 있는 헤더 행 번호 반환."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        for cell in row:
            if cell.value and "Scenario ID" in str(cell.value):
                return i
    return None


def _list_excel_sheets(filepath: Path) -> list:
    """엑셀 파일의 시트별 테스트케이스 수 반환."""
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        header_row = _detect_header_row(ws)
        if header_row is None:
            continue
        count = 0
        for row in ws.iter_rows(min_row=header_row + 2, values_only=False):
            cell_b = row[1].value if len(row) > 1 else None
            if cell_b and "_" in str(cell_b):
                count += 1
        if count > 0:
            sheets.append({"name": name, "count": count})
    wb.close()
    return sheets


def _parse_excel_sheet(wb, sheet_name: str) -> list:
    """엑셀 시트에서 테스트케이스 목록 추출."""
    ws = wb[sheet_name]
    header_row = _detect_header_row(ws)
    if header_row is None:
        return []

    last_main = ""
    last_sub = ""
    cases = []
    for row in ws.iter_rows(min_row=header_row + 2, values_only=False):
        tc_id = row[1].value if len(row) > 1 else None
        if not tc_id or "_" not in str(tc_id):
            continue
        main = str(row[2].value).strip() if len(row) > 2 and row[2].value else ""
        sub = str(row[3].value).strip() if len(row) > 3 and row[3].value else ""
        detail = str(row[4].value).strip() if len(row) > 4 and row[4].value else ""
        summary = str(row[5].value).strip() if len(row) > 5 and row[5].value else ""
        precond = str(row[6].value).strip() if len(row) > 6 and row[6].value else ""
        steps = str(row[7].value).strip() if len(row) > 7 and row[7].value else ""
        expected = str(row[8].value).strip() if len(row) > 8 and row[8].value else ""
        level = str(row[9].value).strip() if len(row) > 9 and row[9].value else ""
        if main:
            last_main = main
        else:
            main = last_main
        if sub:
            last_sub = sub
        else:
            sub = last_sub
        cases.append({
            "main": main, "sub": sub, "detail": detail,
            "summary": summary, "precondition": precond,
            "steps": steps, "expected": expected, "level": level,
        })
    return cases


def _level_to_priority(level: str) -> str:
    level = level.strip()
    if level in ("BAT", "Level 1"):
        return "high"
    elif level == "Level 2":
        return "medium"
    return "low"


def _to_slug(text: str) -> str:
    """TC Summary → 파일명 슬러그."""
    if not text:
        return "unnamed"
    text = text.replace("\n", " ").strip()
    text = re.sub(r'[/\\:*?"<>|.\[\]()>{},]', '', text)
    text = re.sub(r'\s+', '_', text)
    text = re.sub(r'_+', '_', text).strip('_')
    return text[:60]


def _write_tc_files(cases: list, output_dir: Path) -> int:
    """케이스 목록을 tc_*.md 파일로 생성. 생성 건수 반환."""
    output_dir.mkdir(parents=True, exist_ok=True)
    for i, c in enumerate(cases):
        num = str(i + 1).zfill(3)
        slug = _to_slug(c["summary"])
        filepath = output_dir / f"tc_{num}_{slug}.md"
        priority = _level_to_priority(c.get("level", ""))
        tags = []
        if c["main"]:
            clean = re.sub(r'\(.*?\)', '', c["main"]).strip().replace('\n', '')
            if clean:
                tags.append(clean)
        if c["sub"]:
            tags.append(c["sub"].replace('\n', ''))
        if not tags:
            tags = ["general"]
        tags_str = ", ".join(tags)
        steps_lines = [s.strip() for s in c["steps"].split("\n")
                       if s.strip() and not s.strip().startswith("0.")]
        steps_text = "\n".join(steps_lines) if steps_lines else "1. (스텝 미기재)"
        exp_lines = []
        for e in c["expected"].split("\n"):
            e = e.strip()
            if e:
                if not e.startswith("-") and not e.startswith("*"):
                    e = f"- {e}"
                exp_lines.append(e)
        expected_text = "\n".join(exp_lines) if exp_lines else "- (기대결과 미기재)"
        pre_lines = [p.strip() for p in c["precondition"].split("\n") if p.strip()]
        precond_text = "\n".join(pre_lines) if pre_lines else "- 없음"
        title = c["summary"].replace("\n", " ").strip()
        content = (
            f"---\nid: tc_{num}\ndata_key: null\npriority: {priority}\n"
            f"tags: [{tags_str}]\ntype: structured\n---\n"
            f"# {title}\n\n## Precondition\n{precond_text}\n\n"
            f"## Steps\n{steps_text}\n\n## Expected\n{expected_text}\n"
        )
        filepath.write_text(content, encoding="utf-8")
    return len(cases)


class DashboardHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        path = self.path.split("?")[0]
        if path in ("/", "/index.html"):
            self._serve_file(HERE / "index.html", "text/html; charset=utf-8")
        elif path == "/api/dialogs":
            payload = build_dialogs()
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/events":
            self._serve_sse()
        elif path == "/api/dialog":
            self._serve_json(DIALOG_PATH)
        elif path == "/api/state":
            self._serve_json(STATE_PATH)
        elif path == "/api/pages":
            payload = {"pages": list_pages(), "groups": list_testcase_groups()}
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/pipeline_state":
            payload = build_pipeline_state()
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/batch_state":
            payload = build_batch_state()
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/quick_state":
            payload = load_json(QUICK_STATE_PATH) or {}
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/generated_groups":
            payload = list_generated_groups()
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/reports":
            payload = list_reports()
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path.startswith("/reports/"):
            fname = path[len("/reports/"):]
            fpath = REPORTS_DIR / fname
            if fpath.exists() and fpath.suffix == ".html":
                self._serve_file(fpath, "text/html; charset=utf-8")
            else:
                self.send_response(404)
                self.end_headers()

        # ── Import API (GET) ──────────────────────────────────────
        elif path == "/api/import/files":
            payload = {"ok": True, "files": _list_import_files()}
            self._serve_bytes(
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        elif path == "/api/import/sheets":
            from urllib.parse import urlparse, parse_qs
            qs = parse_qs(urlparse(self.path).query)
            fname = qs.get("file", [""])[0]
            if not fname or ".." in fname:
                self._serve_bytes(
                    b'{"ok":false,"error":"file parameter required"}',
                    "application/json; charset=utf-8")
                return
            fpath = IMPORT_DIR / fname
            if not fpath.exists():
                self._serve_bytes(
                    json.dumps({"ok": False, "error": f"{fname} not found"},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
                return
            try:
                sheets = _list_excel_sheets(fpath)
                self._serve_bytes(
                    json.dumps({"ok": True, "sheets": sheets},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
            except Exception as e:
                self._serve_bytes(
                    json.dumps({"ok": False, "error": str(e)},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")

        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        path = self.path.split("?")[0]
        try:
            self._handle_post(path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            msg = json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False).encode("utf-8")
            try:
                self._serve_bytes(msg, "application/json; charset=utf-8")
            except Exception:
                pass

    def _handle_post(self, path):
        # ── 대화 초기화 ──────────────────────────────────────────
        if path == "/api/reset":
            empty = {"pipeline_url": "", "started_at": "", "sessions": []}
            DIALOG_PATH.write_text(
                json.dumps(empty, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        # ── 토론 시작 ─────────────────────────────────────────────
        elif path == "/api/discuss/start":
            import datetime
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8"))
            topic = body.get("topic", "").strip()
            if not topic:
                self._serve_bytes(b'{"ok":false,"error":"topic required"}',
                                  "application/json; charset=utf-8")
                return

            history = []
            if DISCUSS_PATH.exists():
                try:
                    prev = json.loads(DISCUSS_PATH.read_text(encoding="utf-8"))
                    history = prev.get("history", [])
                    if prev.get("step") in ("approved", "rejected", "discussed"):
                        history.append({k: v for k, v in prev.items() if k != "history"})
                except Exception:
                    pass

            discuss = {
                "topic": topic, "step": "pending", "conclusion": "",
                "rejection_reason": "", "rejection_count": 0,
                "created_at": datetime.datetime.now().isoformat(),
                "history": history,
            }
            DISCUSS_PATH.write_text(
                json.dumps(discuss, ensure_ascii=False, indent=2), encoding="utf-8"
            )

            # Claude Code UserPromptSubmit 훅(check_pending_discuss.py)이
            # 다음 프롬프트 제출 시 자동으로 토론 시작을 Claude에게 주입한다.
            self._serve_bytes(b'{"ok":true}',
                              "application/json; charset=utf-8")

        # ── 항목별 투표 ───────────────────────────────────────────
        elif path == "/api/discuss/vote_item":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8"))
            item_id = int(body.get("item_id", -1))
            vote    = body.get("vote", "")  # "approve" | "reject"

            if not DISCUSS_PATH.exists():
                self._serve_bytes(
                    json.dumps({"ok": False, "error": "state/discuss.json 없음"}, ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
                return

            discuss = json.loads(DISCUSS_PATH.read_text(encoding="utf-8"))
            items = discuss.get("conclusion_items", [])
            for item in items:
                if item["id"] == item_id:
                    item["status"] = "approved" if vote == "approve" else "rejected"
                    break
            discuss["conclusion_items"] = items

            all_voted = bool(items) and all(i["status"] != "pending" for i in items)
            if all_voted:
                finalize_team_notes(discuss)
                discuss["step"] = "approved"

            DISCUSS_PATH.write_text(
                json.dumps(discuss, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._serve_bytes(
                json.dumps({"ok": True, "all_voted": all_voted}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )

        # ── run_qa.py 실행 ──────────────────────────────────────────
        elif path == "/api/run_qa":
            import subprocess as sp
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            url = body.get("url", "").strip()
            cases_dir = body.get("cases_dir", "").strip()  # e.g. "login"
            if not url or not cases_dir:
                self._serve_bytes(
                    b'{"ok":false,"error":"url and cases_dir required"}',
                    "application/json; charset=utf-8")
                return
            cases_path = TESTCASES_DIR / cases_dir
            if not cases_path.exists():
                self._serve_bytes(
                    json.dumps({"ok": False, "error": f"testcases/{cases_dir} not found"}, ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
                return
            log_path = LOGS_DIR / "run_qa.txt"
            script = PROJECT_ROOT / "run_qa.py"
            log_file = open(log_path, "w", encoding="utf-8")
            proc = sp.Popen(
                [PYTHON_EXE, "-u", str(script),
                 "--url", url, "--cases", str(cases_path)],
                cwd=str(PROJECT_ROOT),
                stdout=log_file, stderr=sp.STDOUT,
            )
            # 자식 프로세스가 fd를 상속했으므로 부모에서 닫아도 안전
            log_file.close()
            print(f"[Dashboard] run_qa.py 실행 (PID: {proc.pid}, URL: {url}, cases: {cases_dir})")
            self._serve_bytes(
                json.dumps({"ok": True, "pid": proc.pid, "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )

        # ── run_qa_parallel.py 실행 ──────────────────────────────────
        elif path == "/api/run_qa_parallel":
            import subprocess as sp
            log_path = LOGS_DIR / "run_parallel.txt"
            script = PROJECT_ROOT / "run_qa_parallel.py"
            log_file = open(log_path, "w", encoding="utf-8")
            proc = sp.Popen(
                [PYTHON_EXE, "-u", str(script)],
                cwd=str(PROJECT_ROOT),
                stdout=log_file, stderr=sp.STDOUT,
            )
            log_file.close()
            print(f"[Dashboard] run_qa_parallel.py 실행 (PID: {proc.pid})")
            self._serve_bytes(
                json.dumps({"ok": True, "pid": proc.pid, "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )

        # ── 실행 로그 조회 (공통) ────────────────────────────────────
        elif path == "/api/run_log":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            log_name = body.get("log", "run_qa.txt")
            # 보안: 상위 디렉토리 접근 방지
            if ".." in log_name or "/" in log_name:
                self._serve_bytes(b'{"ok":false,"log":""}', "application/json; charset=utf-8")
                return
            log_path = LOGS_DIR / log_name
            if log_path.exists():
                content = log_path.read_text(encoding="utf-8", errors="replace")
                self._serve_bytes(
                    json.dumps({"ok": True, "log": content}, ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8"
                )
            else:
                self._serve_bytes(b'{"ok":false,"log":""}', "application/json; charset=utf-8")

        # ── 단일 파이프라인 승인 ─────────────────────────────────────
        elif path == "/api/pipeline/approve":
            if not STATE_PATH.exists():
                self._serve_bytes(b'{"ok":false,"error":"state/pipeline.json not found"}',
                                  "application/json; charset=utf-8")
                return
            state = json.loads(STATE_PATH.read_text(encoding="utf-8"))
            state["approval_status"] = "approved"
            state["step"] = "approved"
            STATE_PATH.write_text(
                json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        # ── 단일 파이프라인 반려 ─────────────────────────────────────
        elif path == "/api/pipeline/reject":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            reason = body.get("reason", "").strip()
            if not STATE_PATH.exists():
                self._serve_bytes(b'{"ok":false,"error":"state/pipeline.json not found"}',
                                  "application/json; charset=utf-8")
                return
            state = json.loads(STATE_PATH.read_text(encoding="utf-8"))
            state["approval_status"] = "rejected"
            state["rejection_reason"] = reason
            state["rejection_count"] = state.get("rejection_count", 0) + 1
            STATE_PATH.write_text(
                json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        # ── 파이프라인 state/pipeline.json 초기화 ──────────────────────────
        elif path == "/api/pipeline/reset":
            init_state = {
                "url": "", "test_cases": [], "step": "init",
                "dom_info": {}, "plan": [],
                "generated_file_path": "tests/generated/test_generated.py",
                "lint_result": {}, "review_summary": "",
                "approval_status": "", "rejection_reason": "",
                "rejection_count": 0, "execution_result": {},
                "heal_count": 0, "heal_context": {}
            }
            STATE_PATH.write_text(
                json.dumps(init_state, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        # ── 병렬 파이프라인 state/parallel.json 초기화 ────────────────────
        elif path == "/api/parallel/reset":
            init_state = {"status": "", "total_count": 0, "targets": []}
            PARALLEL_STATE_PATH.write_text(
                json.dumps(init_state, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            # heal_context도 정리
            heal_ctx = PROJECT_ROOT / "state" / "heal_context.json"
            if heal_ctx.exists():
                heal_ctx.unlink()
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        # ── 빠른 실행 상태 초기화 ───────────────────────────────────
        elif path == "/api/quick/reset":
            if QUICK_STATE_PATH.exists():
                QUICK_STATE_PATH.unlink()
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        # ── 99_merge.py 실행 ─────────────────────────────────────────
        elif path == "/api/run_merge":
            import subprocess as sp
            merge_script = PROJECT_ROOT / "parallel" / "99_merge.py"
            if not merge_script.exists():
                self._serve_bytes(b'{"ok":false,"error":"99_merge.py not found"}',
                                  "application/json; charset=utf-8")
                return
            # 로그 파일로 출력 저장
            log_path = LOGS_DIR / "merge.txt"
            log_file = open(log_path, "w", encoding="utf-8")
            proc = sp.Popen(
                [PYTHON_EXE, "-u", str(merge_script)],
                cwd=str(PROJECT_ROOT),
                stdout=log_file, stderr=sp.STDOUT,
            )
            log_file.close()
            print(f"[Dashboard] 99_merge.py 실행 (PID: {proc.pid}, 로그: {log_path})")
            self._serve_bytes(
                json.dumps({"ok": True, "message": "99_merge.py started", "pid": proc.pid,
                             "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )

        # ── merge 로그 조회 ──────────────────────────────────────────
        elif path == "/api/merge_log":
            log_path = LOGS_DIR / "merge.txt"
            if log_path.exists():
                content = log_path.read_text(encoding="utf-8", errors="replace")
                self._serve_bytes(
                    json.dumps({"ok": True, "log": content}, ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8"
                )
            else:
                self._serve_bytes(b'{"ok":false,"log":""}',
                                  "application/json; charset=utf-8")

        # ── 빠른 실행 (quick run) ──────────────────────────────────
        elif path == "/api/run_quick":
            import subprocess as sp
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            groups = body.get("groups", [])
            if not groups:
                self._serve_bytes(
                    b'{"ok":false,"error":"groups required"}',
                    "application/json; charset=utf-8")
                return
            # 폴더 존재 검증
            missing = [g for g in groups if not (GENERATED_DIR / g).is_dir()]
            if missing:
                msg = json.dumps(
                    {"ok": False, "error": f"존재하지 않는 폴더: {', '.join(missing)}"},
                    ensure_ascii=False).encode("utf-8")
                self._serve_bytes(msg, "application/json; charset=utf-8")
                return
            log_path = LOGS_DIR / "quick_run.txt"
            merge_script = PROJECT_ROOT / "parallel" / "99_merge.py"
            log_file = open(log_path, "w", encoding="utf-8")
            cmd = [PYTHON_EXE, "-u", str(merge_script), "--quick", "--group"] + groups
            proc = sp.Popen(
                cmd, cwd=str(PROJECT_ROOT),
                stdout=log_file, stderr=sp.STDOUT,
            )
            log_file.close()
            print(f"[Dashboard] 빠른 실행 (PID: {proc.pid}, groups: {groups})")
            self._serve_bytes(
                json.dumps({"ok": True, "pid": proc.pid,
                             "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )

        # ── Excel Import 변환 ─────────────────────────────────────
        elif path == "/api/import/convert":
            import openpyxl
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            fname = body.get("file", "").strip()
            sheet_names = body.get("sheets", [])
            if not fname or not sheet_names:
                self._serve_bytes(
                    b'{"ok":false,"error":"file and sheets required"}',
                    "application/json; charset=utf-8")
                return
            fpath = IMPORT_DIR / fname
            if not fpath.exists():
                self._serve_bytes(
                    json.dumps({"ok": False, "error": f"{fname} not found"},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
                return
            try:
                wb = openpyxl.load_workbook(str(fpath), data_only=True)
                results = []
                for sn in sheet_names:
                    if sn not in wb.sheetnames:
                        results.append({"sheet": sn, "count": 0, "error": "시트 없음"})
                        continue
                    cases = _parse_excel_sheet(wb, sn)
                    folder_name = re.sub(r'\s+', '_', sn.strip().lower())
                    out_dir = TESTCASES_DIR / folder_name
                    # 기존 파일 정리
                    if out_dir.exists():
                        for old in out_dir.glob("tc_*.md"):
                            old.unlink()
                    count = _write_tc_files(cases, out_dir)
                    results.append({"sheet": sn, "count": count,
                                    "folder": f"testcases/{folder_name}/"})
                wb.close()
                self._serve_bytes(
                    json.dumps({"ok": True, "results": results},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
            except Exception as e:
                self._serve_bytes(
                    json.dumps({"ok": False, "error": str(e)},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")

        # ── 전체 반려 ─────────────────────────────────────────────
        elif path == "/api/discuss/reject":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            reason = body.get("reason", "").strip()
            if not DISCUSS_PATH.exists():
                self._serve_bytes(
                    json.dumps({"ok": False, "error": "state/discuss.json 없음"}, ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")
                return
            discuss = json.loads(DISCUSS_PATH.read_text(encoding="utf-8"))
            discuss["step"] = "rejected"
            discuss["rejection_reason"] = reason
            discuss["rejection_count"] = discuss.get("rejection_count", 0) + 1
            DISCUSS_PATH.write_text(
                json.dumps(discuss, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

        else:
            self.send_response(404)
            self.end_headers()

    def _serve_sse(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("X-Accel-Buffering", "no")
        self.end_headers()

        q: queue.Queue = queue.Queue(maxsize=10)
        with _sse_lock:
            _sse_clients.append(q)

        def send(data: str):
            self.wfile.write(f"data: {data}\n\n".encode("utf-8"))
            self.wfile.flush()

        try:
            # 연결 즉시 현재 상태 전송
            send(json.dumps(build_dialogs(), ensure_ascii=False))
            while True:
                try:
                    q.get(timeout=15)
                    send(json.dumps(build_dialogs(), ensure_ascii=False))
                except queue.Empty:
                    # keepalive
                    self.wfile.write(b": ping\n\n")
                    self.wfile.flush()
        except Exception:
            pass
        finally:
            with _sse_lock:
                if q in _sse_clients:
                    _sse_clients.remove(q)

    def _serve_file(self, path: Path, content_type: str):
        if path.exists():
            content = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        else:
            self.send_response(404)
            self.end_headers()

    def _serve_json(self, path: Path):
        content = path.read_bytes() if path.exists() else b'{"pipeline_url":"","started_at":"","sessions":[]}'
        self._serve_bytes(content, "application/json; charset=utf-8")

    def _serve_bytes(self, content: bytes, content_type: str):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def log_message(self, format, *args):
        print(f"[{self.log_date_time_string()}] {format % args}")


def main():
    server = ThreadingHTTPServer(("localhost", PORT), DashboardHandler)
    url = f"http://localhost:{PORT}"
    print(f"[Dashboard] 서버 시작: {url}")
    print(f"[Dashboard] 종료: Ctrl+C")
    webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[Dashboard] 서버 종료")


if __name__ == "__main__":
    main()
